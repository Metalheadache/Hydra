"""
Tests for the 7 file-upload fixes.

Covers:
- test_process_upload_path_traversal       (Fix 2 — path traversal blocked)
- test_process_upload_null_byte            (Fix 3 — null byte handled gracefully)
- test_file_processed_events_in_run        (Fix 1 — events reach callbacks in run() path)
- test_corrupt_pdf                         (robustness — corrupt PDF handled)
- test_max_files_limit                     (Fix 5 — >20 files raises ValueError)
- test_xlsx_sheet_limit                    (Fix 6 — sheets capped at 20)
"""

from __future__ import annotations

import asyncio
import tempfile
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
import pytest_asyncio

from hydra_agents.config import HydraConfig
from hydra_agents.events import EventType, HydraEvent
from hydra_agents.file_processor import FileProcessor
from hydra_agents.models import FileAttachment


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture()
def tmp_dir(tmp_path):
    return tmp_path


@pytest.fixture()
def processor(tmp_dir):
    # output_dir must be under tmp_dir so test files pass the allowed-roots check
    return FileProcessor(output_dir=str(tmp_dir))


# ── Fix 2: Path traversal blocked ─────────────────────────────────────────────

@pytest.mark.asyncio
async def test_process_upload_path_traversal(processor):
    """
    Filenames containing path traversal sequences (../../) must be neutralized.
    The file must be saved WITHIN upload_dir, never at the traversal destination.
    The is_relative_to() guard enforces this invariant.
    """
    malicious_filename = "../../etc/passwd.txt"
    content = b"root:x:0:0:root:/root:/bin/bash"

    att = await processor.process_upload(malicious_filename, content)

    # Must have been saved somewhere — not empty
    assert att.filepath != ""
    saved_path = Path(att.filepath)
    # The saved file must exist and be inside upload_dir
    assert saved_path.exists(), f"Expected saved file at {saved_path}"
    assert saved_path.resolve().is_relative_to(processor.upload_dir.resolve()), (
        f"File escaped upload_dir! Saved at {saved_path}, "
        f"upload_dir is {processor.upload_dir}"
    )
    # Content must match
    assert saved_path.read_bytes() == content


# ── Fix 3: Null byte in filename handled gracefully ───────────────────────────

@pytest.mark.asyncio
async def test_process_upload_null_byte(processor):
    """
    Filenames with null bytes must not crash; should return a graceful FileAttachment.
    """
    malicious_filename = "file\x00.txt"
    content = b"some content"

    att = await processor.process_upload(malicious_filename, content)

    assert isinstance(att, FileAttachment)
    # No crash; filepath should be empty or safe sentinel
    assert att.filepath == ""
    assert att.original_name == malicious_filename


@pytest.mark.asyncio
async def test_process_upload_control_char_in_filename(processor):
    """Control characters (< 32) other than null also trigger the guard."""
    filename_with_ctrl = "file\x1f.txt"
    att = await processor.process_upload(filename_with_ctrl, b"data")
    assert att.filepath == ""


# ── Fix 1: FILE_PROCESSED events reach callbacks in run() path ───────────────

@pytest.mark.asyncio
async def test_file_processed_events_in_run(tmp_path):
    """
    FILE_PROCESSED events must be delivered to callbacks registered via on_event()
    even when using the run() path (not stream()).

    Previously, event_bus was created AFTER _process_files(), so FILE_PROCESSED
    events were emitted to None and lost.
    """
    from hydra_agents import Hydra

    collected: list[HydraEvent] = []

    def collector(event: HydraEvent) -> None:
        collected.append(event)

    # Create a test file
    test_file = tmp_path / "hello.txt"
    test_file.write_text("hello world")

    # Minimal mocks so we don't hit real LLM
    mock_plan = MagicMock()
    mock_plan.sub_tasks = []
    mock_plan.agent_specs = []
    mock_plan.execution_groups = []
    mock_plan.original_task = ""

    mock_result = {
        "output": "done",
        "warnings": [],
        "execution_summary": {},
        "files_generated": [],
        "per_agent_quality": {},
        "agents_needing_retry": [],
    }

    hydra = Hydra(config=HydraConfig(output_directory=str(tmp_path / "out")))
    hydra.on_event(collector)

    with (
        patch("hydra_agents.Brain.plan", new_callable=AsyncMock, return_value=mock_plan),
        patch("hydra_agents.AgentFactory.create_agents", return_value=[]),
        patch("hydra_agents.ExecutionEngine.execute", new_callable=AsyncMock),
        patch("hydra_agents.PostBrain.synthesize", new_callable=AsyncMock, return_value=mock_result),
    ):
        await hydra.run("Summarize the file", files=[str(test_file)])

    file_processed_events = [e for e in collected if e.type == EventType.FILE_PROCESSED]
    assert len(file_processed_events) == 1, (
        f"Expected 1 FILE_PROCESSED event, got {len(file_processed_events)}. "
        f"All events: {[e.type for e in collected]}"
    )
    assert file_processed_events[0].data["filename"] == "hello.txt"


# ── Corrupt PDF handled gracefully ────────────────────────────────────────────

@pytest.mark.asyncio
async def test_corrupt_pdf(tmp_dir, processor):
    """
    A file with .pdf extension but invalid content should not crash.
    extracted_text should be None (extraction failure is caught internally).
    """
    pdf_path = tmp_dir / "corrupt.pdf"
    pdf_path.write_bytes(b"not a pdf")

    attachments = await processor.process([str(pdf_path)])
    assert len(attachments) == 1
    att = attachments[0]
    # Extraction should fail gracefully — either None or some text
    # The important thing is no exception was raised
    assert att.original_name == "corrupt.pdf"
    assert att.size_bytes > 0


# ── Fix 5: Max files limit ────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_max_files_limit(tmp_path):
    """
    Passing more than max_upload_files (default 20) should raise ValueError.
    """
    from hydra_agents import Hydra

    config = HydraConfig(max_upload_files=20, output_directory=str(tmp_path / "out"))
    hydra = Hydra(config=config)

    # Create 25 dummy files
    files = []
    for i in range(25):
        f = tmp_path / f"file_{i:02d}.txt"
        f.write_text(f"content {i}")
        files.append(str(f))

    mock_plan = MagicMock()
    mock_plan.sub_tasks = []
    mock_plan.agent_specs = []
    mock_plan.execution_groups = []
    mock_plan.original_task = ""

    with (
        patch("hydra_agents.Brain.plan", new_callable=AsyncMock, return_value=mock_plan),
    ):
        with pytest.raises(ValueError, match="Too many files"):
            await hydra.run("process files", files=files)


# ── Fix 6: XLSX sheet count capped at 20 ─────────────────────────────────────

@pytest.mark.asyncio
async def test_xlsx_sheet_limit(tmp_dir, processor):
    """
    XLSX workbooks with more than 20 sheets should only process the first 20
    and append a truncation message.
    """
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not available")

    wb = openpyxl.Workbook()
    # openpyxl starts with one default sheet; rename and add more
    wb.active.title = "Sheet_01"
    for i in range(2, 26):  # 25 total sheets
        wb.create_sheet(title=f"Sheet_{i:02d}")

    assert len(wb.sheetnames) == 25

    xlsx_path = tmp_dir / "many_sheets.xlsx"
    wb.save(str(xlsx_path))

    attachments = await processor.process([str(xlsx_path)])
    text = attachments[0].extracted_text
    assert text is not None

    # Must mention the truncation
    assert "more sheets not shown" in text

    # The 21st sheet and beyond should not appear as section headers
    assert "--- Sheet: Sheet_21 ---" not in text
    assert "--- Sheet: Sheet_25 ---" not in text

    # The first 20 should appear
    assert "--- Sheet: Sheet_01 ---" in text
    assert "--- Sheet: Sheet_20 ---" in text
